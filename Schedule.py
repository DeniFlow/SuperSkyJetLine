
import openpyxl
from datetime import date
from enum import Enum

class Weekday(Enum):
    Понедельник = 1
    Вторник = 2
    Среда = 3
    Четверг = 4
    Пятница = 5
    Суббота = 6
    @classmethod
    def from_date(cls, date):
        return cls(date.isoweekday())
class Schedule:
    __path = 'schedule.xlsx'
    __wb_obj = openpyxl.load_workbook(__path)
    __sheet_obj = __wb_obj.active
    __row = __sheet_obj.max_row
    __column = __sheet_obj.max_column
    def GetShedule(self,id,week):
        daysOfWeek = []
        schedule = ["Время начала пар:",
                    "1)8.00 - 9.35 " + "2)9.45 - 11.20 " + "3)11.50 - 13.25 " + "4)13.35 - 15.10 " + "5)15.40 - 17.15 " + "6)17.25 - 19.00"]
        endEvenWeek = self.__row//2+1
        startOddWeek = self.__row//2+3
        for i in range(1, self.__column + 1):
            days = self.__sheet_obj.cell(row=1, column=i)
            daysOfWeek += [days.value]
        if week == 'Чётная':
            for j in range(1, self.__column + 1):
                lessons = []
                for i in range(2, endEvenWeek):
                    cell_obj = self.__sheet_obj.cell(row=i, column=j)
                    if cell_obj.value != None:
                        lessons += [cell_obj.value]
                    else:
                        lessons += ['---']
                schedule += [{daysOfWeek[j - 1]: lessons}]
            return schedule
        if week == 'Нечётная':
            for j in range(1, self.__column + 1):
                lessons = []
                for i in range(startOddWeek, self.__row + 1):
                    cell_obj = self.__sheet_obj.cell(row=i, column=j)
                    if cell_obj.value != None:
                        lessons += [cell_obj.value]
                    else:
                        lessons += ['---']
                schedule += [{daysOfWeek[j - 1]: lessons}]
            return schedule

    def PrintSchedule(self, id, week):
        week = self.GetShedule(id, week)
        for i in range(len(week)):
            print(week[i],end = '\n')
    def PrintDay(self, id, week, day):
        week = self.GetShedule(id, week)
        return week[day]